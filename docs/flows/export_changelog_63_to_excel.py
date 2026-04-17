from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


DOC_PATH = Path(__file__).with_name("bang-pha-buoc-event-e2e.md")
OUT_PATH = Path(__file__).with_name("bang-pha-buoc-event-e2e_6.3_changelog.xlsx")


def strip_markdown(text: str) -> str:
    """Rút gọn markdown để dễ sửa trong Excel."""
    out = text
    out = re.sub(r"\*\*(.*?)\*\*", r"\1", out)
    out = re.sub(r"`([^`]*)`", r"\1", out)
    out = re.sub(r"\[([^\]]+)\]\(([^)]+)\)", r"\1 (\2)", out)
    return out.strip()


def parse_changelog_section(lines: list[str]) -> list[tuple[int, str, str]]:
    in_section = False
    rows: list[tuple[int, str, str]] = []
    item_index = 0

    for raw_line in lines:
        line = raw_line.rstrip("\n")
        if line.startswith("### 6.3. Changelog"):
            in_section = True
            continue
        if not in_section:
            continue
        if line.startswith("### "):
            break
        if not line.startswith("- "):
            continue

        content = line[2:].strip()
        m = re.match(r"\*\*(\d{4}-\d{2}-\d{2}):\*\*\s*(.*)", content)
        if m:
            date_str = m.group(1)
            detail = m.group(2).strip()
        else:
            date_str = ""
            detail = content

        item_index += 1
        rows.append((item_index, date_str, strip_markdown(detail)))

    return rows


def export_to_xlsx(rows: list[tuple[int, str, str]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "6.3 Changelog"

    ws.append(["STT", "Ngày", "Nội dung"])
    for row in rows:
        ws.append(list(row))

    header_font = Font(bold=True)
    for c in ("A1", "B1", "C1"):
        ws[c].font = header_font
        ws[c].alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 180
    ws.freeze_panes = "A2"

    for row_idx in range(2, ws.max_row + 1):
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="center", vertical="top")
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="center", vertical="top")
        ws[f"C{row_idx}"].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)


def main() -> None:
    lines = DOC_PATH.read_text(encoding="utf-8").splitlines(keepends=True)
    rows = parse_changelog_section(lines)
    if not rows:
        raise RuntimeError("Không tìm thấy dữ liệu trong mục 6.3 Changelog.")
    export_to_xlsx(rows, OUT_PATH)
    print(f"Đã xuất {len(rows)} dòng -> {OUT_PATH}")


if __name__ == "__main__":
    main()
