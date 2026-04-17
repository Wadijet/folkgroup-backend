from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


DOC_PATH = Path(__file__).with_name("bang-pha-buoc-event-e2e.md")
OUT_PATH = Path(__file__).with_name("bang-pha-buoc-event-e2e_5.3_bang-chi-tiet.xlsx")


def split_markdown_row(line: str) -> list[str]:
    s = line.strip()
    if not s.startswith("|"):
        return []
    s = s.strip("|")
    return [part.strip() for part in s.split("|")]


def strip_markdown(text: str) -> str:
    out = text
    out = re.sub(r"\*\*(.*?)\*\*", r"\1", out)
    out = re.sub(r"`([^`]*)`", r"\1", out)
    out = re.sub(r"\[([^\]]+)\]\(([^)]+)\)", r"\1 (\2)", out)
    return out.strip()


def is_separator_row(cells: list[str]) -> bool:
    if not cells:
        return False
    for c in cells:
        t = c.replace(" ", "")
        if t and not set(t) <= {"-"}:
            return False
    return True


def parse_table_53(lines: list[str]) -> tuple[list[str], list[list[str]]]:
    table_start = -1
    for i, line in enumerate(lines):
        if line.startswith("| Giai đoạn | Bước | Sự kiện |"):
            table_start = i
            break
    if table_start < 0:
        raise RuntimeError("Không tìm thấy header bảng §5.3.")

    rows: list[list[str]] = []
    header: list[str] | None = None

    for line in lines[table_start:]:
        if line.startswith("#### Bổ sung: `eventType` domain"):
            break
        if not line.startswith("|"):
            if header is not None:
                break
            continue

        cells = split_markdown_row(line)
        if not cells:
            continue
        if is_separator_row(cells):
            continue

        if header is None:
            header = [strip_markdown(c) for c in cells]
            continue

        if len(cells) != len(header):
            raise RuntimeError(
                f"Số cột không khớp ở bảng §5.3: expected={len(header)} got={len(cells)}"
            )
        rows.append([strip_markdown(c) for c in cells])

    if header is None or not rows:
        raise RuntimeError("Không đọc được dữ liệu bảng §5.3.")
    return header, rows


def export_xlsx(header: list[str], rows: list[list[str]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bang 5.3"

    ws.append(header)
    for row in rows:
        ws.append(row)

    # Header style
    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Width
    width_map = {
        1: 10,   # Giai đoạn
        2: 12,   # Bước
        3: 14,   # Sự kiện
        4: 100,  # Mô tả kỹ thuật
        5: 80,   # Mô tả người dùng
        6: 45,   # eventType
        7: 30,   # eventSource
        8: 28,   # pipelineStage
        9: 40,   # module giao việc
        10: 34,  # module thực hiện
        11: 20,  # nhóm trách nhiệm
    }
    for col_idx, width in width_map.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Data style
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(header) + 1):
            align = Alignment(vertical="top", wrap_text=True)
            if c in (1, 2, 3, 11):
                align = Alignment(horizontal="center", vertical="top", wrap_text=True)
            ws.cell(row=r, column=c).alignment = align

    ws.freeze_panes = "A2"
    wb.save(out_path)


def main() -> None:
    lines = DOC_PATH.read_text(encoding="utf-8").splitlines()
    header, rows = parse_table_53(lines)
    export_xlsx(header, rows, OUT_PATH)
    print(f"Đã xuất {len(rows)} dòng -> {OUT_PATH}")


if __name__ == "__main__":
    main()
